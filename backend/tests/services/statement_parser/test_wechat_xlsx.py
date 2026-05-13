"""微信支付 xlsx 解析器测试。"""
from datetime import datetime
from decimal import Decimal

import pytest

from app.services.statement_parser.wechat_xlsx import WechatXlsxParser


@pytest.fixture(scope="module")
def parser() -> WechatXlsxParser:
    return WechatXlsxParser()


@pytest.fixture(scope="module")
def parsed(parser: WechatXlsxParser, wechat_xlsx_bytes: bytes):
    return parser.parse(wechat_xlsx_bytes)


def test_source_type(parser: WechatXlsxParser):
    assert parser.source_type == "wechat_xlsx"


def test_detect_accepts_real_wechat_xlsx(
    parser: WechatXlsxParser, wechat_xlsx_bytes: bytes, wechat_filename: str
):
    assert parser.detect(wechat_xlsx_bytes, wechat_filename) is True


def test_detect_rejects_csv(parser: WechatXlsxParser):
    assert parser.detect(b"a,b,c\n1,2,3", "x.csv") is False


def test_detect_rejects_unrelated_xlsx(parser: WechatXlsxParser):
    """简单的 xlsx 文件头但内容不含微信特征,应被拒。"""
    # ZIP 文件头(xlsx 是 zip)
    fake_xlsx = b"PK\x03\x04" + b"\x00" * 100
    assert parser.detect(fake_xlsx, "other.xlsx") is False


def test_parse_yields_transactions(parsed):
    assert len(parsed.raw_transactions) >= 1


def test_parse_account_hint_is_wechat(parsed):
    h = parsed.account_hint
    assert h.type == "wechat"
    assert h.institution == "微信支付"
    assert h.last4 is None  # 微信全局账户,具体卡在 payment_method_raw


def test_parse_amounts_are_positive_decimals(parsed):
    for tx in parsed.raw_transactions:
        assert isinstance(tx.amount, Decimal)
        assert tx.amount > 0
        assert tx.currency == "CNY"
        assert tx.amount_settled_cny == tx.amount


def test_parse_tx_kind_three_states(parsed):
    """收/支 列三态:支出 / 收入 / 中性交易。"""
    kinds = {tx.tx_kind for tx in parsed.raw_transactions}
    assert kinds.issubset({"expense", "income", "neutral"})
    assert "expense" in kinds  # 真实样本必有支出


def test_parse_payment_method_raw_populated_for_card_payments(parsed):
    """有"建设银行信用卡(7432)"等的行,payment_method_raw 必填且含末 4 位。"""
    card_method_txs = [
        tx for tx in parsed.raw_transactions
        if tx.payment_method_raw and any(
            kw in tx.payment_method_raw for kw in ["银行", "信用卡", "储蓄卡"]
        )
    ]
    assert len(card_method_txs) >= 1, "样本中至少应有一条用银行卡支付的微信交易"
    # 抽样检查一条:payment_method_raw 必含 4 位数字
    import re
    for tx in card_method_txs[:3]:
        assert re.search(r"\d{4}", tx.payment_method_raw), \
            f"payment_method_raw 应含末 4 位卡号: {tx.payment_method_raw}"


def test_parse_external_tx_id_populated(parsed):
    """微信交易单号入 external_tx_id。"""
    has_id = sum(1 for tx in parsed.raw_transactions if tx.external_tx_id)
    # 不强求 100%(中性交易如零钱通转入可能没有),但绝大多数应有
    assert has_id >= len(parsed.raw_transactions) * 0.8


def test_parse_metadata_counts(parsed):
    md = parsed.metadata
    assert "raw_row_count" in md
    assert "imported_count" in md
    assert md["imported_count"] == len(parsed.raw_transactions)


def test_parse_period_covers_2025_q4_to_2026_q1(parsed):
    """样本期 20251226-20260326,跨年。"""
    assert parsed.period_start.year in (2025, 2026)
    assert parsed.period_end.year in (2025, 2026)
    assert parsed.period_start <= parsed.period_end


def test_parse_invalid_bytes_raises(parser: WechatXlsxParser):
    with pytest.raises(ValueError):
        parser.parse(b"not an xlsx at all")


# ── helper 函数单元测试(覆盖 missing 行) ─────────────────────────────────────

from app.services.statement_parser.wechat_xlsx import (
    _parse_amount,
    _to_str,
    _infer_tx_kind,
)


class TestParseAmount:
    """_parse_amount 各路径覆盖。"""

    def test_none_returns_zero(self):
        assert _parse_amount(None) == Decimal("0")

    def test_int_value(self):
        assert _parse_amount(10) == Decimal("10")

    def test_float_value(self):
        result = _parse_amount(3.14)
        assert result == Decimal("3.14")

    def test_string_with_yen_symbol(self):
        assert _parse_amount("¥123.45") == Decimal("123.45")

    def test_string_with_xa5(self):
        assert _parse_amount("\xa5999.00") == Decimal("999.00")

    def test_string_with_comma(self):
        assert _parse_amount("1,234.56") == Decimal("1234.56")

    def test_empty_string_returns_zero(self):
        assert _parse_amount("   ") == Decimal("0")

    def test_invalid_string_returns_zero(self):
        assert _parse_amount("abc") == Decimal("0")


class TestToStr:
    """_to_str 各路径覆盖。"""

    def test_none_returns_empty(self):
        assert _to_str(None) == ""

    def test_slash_returns_empty(self):
        assert _to_str("/") == ""

    def test_slash_with_surrounding_whitespace_returns_empty(self):
        """微信账单偶有 " / " 这种带空白的占位,strip 后视为占位。"""
        assert _to_str(" / ") == ""

    def test_normal_string(self):
        assert _to_str("hello") == "hello"

    def test_strips_whitespace(self):
        assert _to_str("  hello  ") == "hello"

    # 回归 B-poly-3:含 "/" 但非纯占位的商户名必须原样保留
    def test_merchant_with_internal_slash_preserved(self):
        assert _to_str("A/B 公司") == "A/B 公司"

    def test_merchant_starts_with_slash_preserved(self):
        assert _to_str("/start") == "/start"

    def test_merchant_ends_with_slash_preserved(self):
        assert _to_str("end/") == "end/"

    def test_short_two_char_with_slash_preserved(self):
        assert _to_str("a/b") == "a/b"


class TestInferTxKind:
    """_infer_tx_kind 各路径覆盖。"""

    def test_expense(self):
        assert _infer_tx_kind("支出") == "expense"

    def test_income(self):
        assert _infer_tx_kind("收入") == "income"

    def test_slash_neutral(self):
        assert _infer_tx_kind("/") == "neutral"

    def test_none_neutral(self):
        assert _infer_tx_kind(None) == "neutral"

    def test_other_neutral(self):
        assert _infer_tx_kind("other") == "neutral"


def test_parse_missing_columns_raises(parser: WechatXlsxParser):
    """xlsx 缺少必需列时应抛 ValueError。"""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    # 第 1 行写微信账单标识(让 detect 通过),第 18 行写不完整表头
    ws.cell(row=1, column=1, value="微信支付账单明细")
    ws.cell(row=18, column=1, value="Wrong Column")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match="missing columns"):
        parser.parse(buf.read())


def test_parse_too_short_xlsx_raises(parser: WechatXlsxParser):
    """行数不够 18 行时应抛 ValueError。"""
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="微信支付账单明细")
    # 只有 5 行,远不够 18 行
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"row{i}")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError):
        parser.parse(buf.read())
