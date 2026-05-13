"""微信支付 xlsx 解析器。

spec § 5.3.2:
- 跳前 17 行元信息,第 18 行(1-based)表头,数据从第 19 行
- 11 列,关键:支付方式列含底层卡 "建设银行信用卡(7432)"
- 收/支:支出 → expense / 收入 → income / 中性交易 → neutral
- 抽末 4 位卡号塞 payment_method_raw,跨源去重(slice C)用

实际样本偏离 plan 的微调(2026-05-09 确认):
  1. "收/支" 列三态为 "支出"/"收入"/"/"(斜杠),而非文字"中性交易"
     — 斜杠统一映射 neutral(与微信账单页眉"中性交易:1笔"一致)
  2. 金额列 openpyxl 直接读成 float/int,无 "¥" 前缀;_parse_amount 先 str() 再 Decimal
  3. 交易时间 openpyxl 直接读成 datetime 对象(Excel 时间单元格自动识别)
  4. "收/支" = "/" 的行 payment_method_raw 可能也是 "/",此时置 None
  5. 样本中无文字"中性交易",全部标"/"——_infer_tx_kind 处理斜杠 → neutral
"""
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.services.statement_parser.base import (
    AccountHint,
    ParseResult,
    RawTransaction,
)


# 元信息特征:微信账单 A1 单元格通常含 "微信支付账单明细" 等
_META_MARKERS = [
    "微信支付账单明细",   # 微信支付账单明细
    "微信账单",                             # 微信账单
    "微信支付商户账单",    # 微信支付商户账单
    "微信支付个人账单",    # 微信支付个人账单
]

# 分隔线特征(第 17 行):含 "微信支付账单明细列表"
_SEPARATOR_MARKER = "微信支付账单明细列表"  # 微信支付账单明细列表

_HEADER_ROW_1BASED = 18  # 第 18 行是表头(spec § 5.3.2,真实样本已验证)

# 斜杠占位符(微信用 "/" 填无意义字段)
_SLASH = "/"


def _parse_amount(val) -> Decimal:
    """微信金额列 openpyxl 读成 float/int;转 Decimal。
    也兼容含 ¥ 符号的字符串(未来兼容)。
    """
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        # float 先转字符串再构造 Decimal,避免浮点精度问题
        return Decimal(str(val))
    s = str(val).strip().replace("\xa5", "").replace("¥", "").replace(",", "")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _to_str(val) -> str:
    """None 和斜杠占位都视作空串。

    重要:仅当 strip 后**整字段就是单个 "/"** 时才视为占位。
    商户名含 "/" 字符(如 "A/B 公司"、"/start"、"end/")必须原样保留 ——
    不要改成 `_SLASH in s` 或类似 contains 判断,否则会吞掉合法商户名。
    """
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s == _SLASH else s


def _infer_tx_kind(in_or_out) -> str:
    """收/支 列 → tx_kind。

    微信账单三种取值:
      "支出" → expense
      "收入" → income
      "/"   → neutral (微信用斜杠表示中性交易,如信用卡还款/零钱通存取)
    """
    s = str(in_or_out).strip() if in_or_out is not None else ""
    if s == "支出":   # 支出
        return "expense"
    if s == "收入":   # 收入
        return "income"
    # "/" 或 "中性交易" 或其他均为 neutral
    return "neutral"


class WechatXlsxParser:
    source_type = "wechat_xlsx"

    def detect(self, file_bytes: bytes, filename: str) -> bool:
        """嗅探:.xlsx 扩展名 + A1-A5 单元格含微信账单标识。"""
        if not filename.lower().endswith(".xlsx"):
            return False
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return False
        try:
            ws = wb.active
            # 取 A1-A5 内容,有任何一行命中 _META_MARKERS 即认
            for row_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=1).value
                if cell and isinstance(cell, str):
                    if any(m in cell for m in _META_MARKERS):
                        return True
        except Exception:
            return False
        finally:
            wb.close()
        return False

    def parse(self, file_bytes: bytes) -> ParseResult:
        try:
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"wechat xlsx load failed: {e}") from e

        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if len(rows) < _HEADER_ROW_1BASED:
            raise ValueError(
                f"wechat xlsx too short (expect >= {_HEADER_ROW_1BASED} rows incl. meta + header)"
            )

        # 第 18 行(index 17)是表头
        header = [str(c).strip() if c is not None else "" for c in rows[_HEADER_ROW_1BASED - 1]]
        data_rows = rows[_HEADER_ROW_1BASED:]  # 第 19 行起是数据

        # 关键列校验
        required = [
            "交易时间",   # 交易时间
            "交易类型",   # 交易类型
            "交易对方",   # 交易对方
            "收/支",              # 收/支
            "金额(元)",       # 金额(元)
            "支付方式",   # 支付方式
            "当前状态",   # 当前状态
            "交易单号",   # 交易单号
        ]
        missing = [k for k in required if k not in header]
        if missing:
            raise ValueError(f"wechat xlsx missing columns: {missing}")

        # 构建列索引
        idx = {h: i for i, h in enumerate(header) if h}

        def _get(row, col_name, default=None):
            i = idx.get(col_name)
            if i is None or i >= len(row):
                return default
            return row[i]

        raw_rows: list[tuple] = []
        for row in data_rows:
            if not row or all(c in (None, "") for c in row):
                continue
            raw_rows.append(row)

        txs: list[RawTransaction] = []
        all_times: list[datetime] = []

        for row in raw_rows:
            # 金额
            try:
                amount = _parse_amount(_get(row, "金额(元)"))  # 金额(元)
            except (InvalidOperation, Exception) as e:
                raise ValueError(f"wechat xlsx bad amount in row {row}: {e}") from e

            if amount <= 0:
                continue

            # 交易时间:openpyxl 自动将 Excel 时间单元格解析为 datetime
            tx_time_raw = _get(row, "交易时间")  # 交易时间
            if isinstance(tx_time_raw, datetime):
                tx_time = tx_time_raw
            elif tx_time_raw is not None:
                s = str(tx_time_raw).strip()
                if not s:
                    continue
                try:
                    tx_time = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
            else:
                continue
            all_times.append(tx_time)

            # 收/支
            tx_kind = _infer_tx_kind(_get(row, "收/支"))  # 收/支

            # 交易对方
            merchant_raw = _to_str(_get(row, "交易对方"))  # 交易对方

            # 商品描述:商品列,fallback 交易类型
            desc_raw = _to_str(_get(row, "商品")) or _to_str(_get(row, "交易类型"))  # 商品 / 交易类型

            # 交易单号
            ext_tx_id_raw = _to_str(_get(row, "交易单号"))  # 交易单号
            # 交易单号为空或全斜杠时 None
            external_tx_id = ext_tx_id_raw if ext_tx_id_raw else None

            # 商户单号
            ext_merchant_raw = _to_str(_get(row, "商户单号"))  # 商户单号
            external_merchant_id = ext_merchant_raw if ext_merchant_raw else None

            # 支付方式:保留原文(含末 4 位卡号),供跨源去重
            pay_method_raw = _to_str(_get(row, "支付方式"))  # 支付方式
            payment_method_raw = pay_method_raw if pay_method_raw else None

            # 原始行 dict(全部列)
            raw_row_dict = {
                header[i]: (str(v) if v is not None else "")
                for i, v in enumerate(row)
                if i < len(header) and header[i]
            }

            txs.append(RawTransaction(
                tx_time=tx_time,
                post_time=None,
                amount=amount,
                currency="CNY",
                amount_settled_cny=amount,
                tx_kind=tx_kind,
                merchant_raw=merchant_raw,
                counterparty_raw=merchant_raw or None,
                description_raw=desc_raw or None,
                external_tx_id=external_tx_id,
                external_merchant_id=external_merchant_id,
                payment_method_raw=payment_method_raw,
                raw_row=raw_row_dict,
            ))

        period_start = min(all_times) if all_times else datetime(1970, 1, 1)
        period_end = max(all_times) if all_times else datetime(1970, 1, 1)

        return ParseResult(
            raw_transactions=txs,
            account_hint=AccountHint(
                type="wechat",
                institution="微信支付",  # 微信支付
                last4=None,
            ),
            period_start=period_start,
            period_end=period_end,
            metadata={
                "raw_row_count": len(raw_rows),
                "imported_count": len(txs),
                "dropped_count": len(raw_rows) - len(txs),
            },
        )
